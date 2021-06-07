#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
対象とするユーザIDを入力すると以下の動作を実行する
1) 対象ユーザがOrG内で所属するチャンネル（パブリック/プライベート/DM/グループDM）を一覧にして output/ に出力する
2) 1)で出力された情報をもとに、 export/ ディレクトリ配下にあるSlackのエクスポートデータから、対象チャンネルのログを取得し成形し、output/ に出力する
"""
import os, sys
import csv
import urllib.request, urllib.error, urllib.parse
import json
import configparser
import time, datetime
import shutil
import glob
import openpyxl, openpyxl.styles


class Exec_api:
    def exec (self, req):
        """
        explanation:
            exec Slack API
        Args:
            req: urllib request object
        Return:
            body: Json object (dict)
        正常に完了した場合は Responsbody(json) を返す
        失敗した場合は、エラーjson(dict) を返す
        {"ok": false, "err":{"code": $err.code, "reason": $err.reason}}
        """
        body = {"ok": False}

        try:
            with urllib.request.urlopen(req) as res:
                body = json.loads(res.read().decode('utf-8'))
        except urllib.error.HTTPError as err:
            time.sleep (61)
            try:
                with urllib.request.urlopen(req) as res:
                    body = json.loads(res.read().decode('utf-8'))
            except urllib.error.HTTPError as err:
                err_d = {'reason': str(err.reason), 'code': str(err.code)}
                body = {'ok': False, 'err':err_d}

        except urllib.error.URLError as err:
            time.sleep (11)
            try:
                with urllib.request.urlopen(req) as res:
                    body = json.loads(res.read().decode('utf-8'))
            except urllib.error.URLError as err:
                err_d = {'reason': str(err.reason)}
                body = {'ok': False, 'err':err_d}

        return body


class Api:
    """
    Slack REST API を実行するための Class
    """
    def d_enterprise_info (self):
        """
        # GET
        https://api.slack.com/enterprise/discovery/methods#enterprise_info
        teamidとワークスペース名の対応を作るのに利用する
        """
        url = "https://slack.com/api/discovery.enterprise.info"
        url = url + "?token=" + token
        req = urllib.request.Request (url)
        api = Exec_api ()
        body = api.exec (req)
        return body


    def d_user_info (self, user:str) -> dict:
        """
        # GET
        https://api.slack.com/enterprise/discovery/methods#user_info
        ユーザIDとユーザ名の対応を作るのに利用する
        """
        url = "https://slack.com/api/discovery.user.info"
        url = url + "?token=" + token + "&user=" + user
        req = urllib.request.Request (url)
        api = Exec_api ()
        body = api.exec (req)
        return body

    def d_user_conv (self, user:str, c_type:str) -> dict:
        """
        # GET
        https://api.slack.com/enterprise/discovery/methods#user_conversations
        ユーザが所属するチャンネル一覧を取得するのに利用する
        Todo: パブリック/プライベート/DM/グループDM それぞれ毎に 500 チャンネル以上所属している場合ページネーションの処理が必要。未実装
        ※そうそういらないと思うので現状実装を見送り
        """
        url = "https://slack.com/api/discovery.user.conversations"
        url = url + "?token=" + token + "&user=" + user + "&include_historical=true" + "&limit=500"
        if c_type == "public":
            url = url + "&only_public=true"
        elif c_type == "private":
            url = url + "&only_private=true"
        elif c_type == "im":
            url = url + "&only_im=true"
        elif c_type == "mpim":
            url = url + "&only_mpim=true"
        else:
            pass
        req = urllib.request.Request (url)
        api = Exec_api ()
        body = api.exec (req)
        return body

    def d_conv_info (self, channel:str, team:str ) -> dict:
        """
        # GET
        https://api.slack.com/enterprise/discovery/methods#conversations_info
        チャンネル名とチャンネルIDの対応に利用する
        ※エクスポートファイルの中の channels.json/groups.json/dms.json/mpims.json を利用すれば不要かも
        """
        url = "https://slack.com/api/discovery.conversations.info"
        url = url + "?token=" + token + "&team=" + team + "&channel=" + channel
        req = urllib.request.Request (url)
        api = Exec_api ()
        body = api.exec (req)
        return body

"""
その他の関数
"""
def loadconf ():
    global token, ORG_ID, ORG_NAME, ORG_DOMAIN, EXPORT_DIR
    cfg = configparser.ConfigParser ()
    cfg_file = os.path.dirname(__file__) + '/config.ini'
    cfg.read (cfg_file)
    token = cfg['slack']['token']
    ORG_ID = cfg['slack']['org_id']
    ORG_NAME = cfg['slack']['org_name']
    ORG_DOMAIN = cfg['slack']['org_domain']
    EXPORT_DIR = cfg['slack']['export_dir']
    return 0

def print_progress_cycle (message:str, i:int):
    if i%4 == 0:
        x = '|'
    elif i%4 == 1:
        x = '/'
    elif i%4 == 2:
        x = '-'
    elif i%4 == 3:
        x = '\\'
    else:
        x = '|'

    print (message + x, '\r', end='')

def get_temas_info ():
    api = Api ()
    res = api.d_enterprise_info ()
    teams = {ORG_ID:[ORG_NAME,ORG_DOMAIN]}
    for data in res['enterprise']['teams']:
        teams[data['id']] = [data['name'], data['domain']]
    return teams


def main():

    now = datetime.datetime.now()
    loadconf ()
    print ("ID: ", end="")
    user = input()

    OUTPUT_DIR = 'output/' + user + '_' + f"{now:%Y%m%d_%H%M%S}"
    os.makedirs (OUTPUT_DIR)

    api = Api ()
    """
    Step1: EG 配下のWS一覧を取得しておく
    """
    print ("STEP1: get OrG teams info...", '\r', end='')
    teams = get_temas_info ()
    print ("STEP1: get OrG teams info... done")

    """
    Step2: 対象ユーザの所属チャンネル一覧を取得する
    """
    print ("STEP2: get channels info target user has joind (include already left)...", '\r', end='')
    channels_csv = OUTPUT_DIR + '/' +user + '_' + f"{now:%Y%m%d_%H%M%S}" + '.csv'
    with open(channels_csv, mode='w', encoding="utf_8_sig") as f:
        writer = csv.writer(f)
        writer.writerow (['channel_id', 'channel_name', 'team_id', 'team_name', 'team_domain', 'type'])

    c_types = ['public', 'private', 'im', 'mpim']
    for c_type in c_types:
        res = api.d_user_conv (user, c_type)
        if not res['ok']:
            print ("STEP2: get channels info target user has joind (include already left)...")
            print (" -> ERROR : " + res['error'], file=sys.stderr)
            exit (1)

        convs = res['channels']
        ret = []

#        いずれ軽量化する時に使う
#        if c_type == 'public':
#            f_name = 'channels.json'
#        elif c_type == 'private':
#            f_name = 'groups.json'
#        elif c_type == 'mpim':
#            f_name = 'mpims.json'
#        else: # c_type == 'im'
#            f_name = 'dms.json' # stab

        x = 0
        if c_type == 'im':
            # DMの場合、チャンネル名がない
            for conv in convs:
                x += 1
                print_progress_cycle ("STEP2: get channels info target user has joind (include already left)...", x)

                c = []
                c.append (conv['id'])
                c.append (conv['id']) # DMはチャンネル名がないため、形式上チャンネルIDを二回いれておく
                c.append (conv['team_id'])
                c.append (teams[conv['team_id']][0]) # チーム名
                c.append (teams[conv['team_id']][1]) # チームドメイン
                c.append (c_type)
                ret.append(c)
            with open(channels_csv, mode='a', encoding="utf_8_sig") as f:
                writer = csv.writer(f)
                for line in ret:
                    writer.writerow (line)

        else:
            # DM 以外の処理
            # チャンネル名は d_convs_info () を使って確認する
            # Todo: エクスポートの channels.json を使って軽くする
            for conv in convs:
                if conv['team_id'] in teams:
                    x += 1
                    print_progress_cycle ("STEP2: get channels info target user has joind (include already left)...", x)
                    c = []
                    id = conv['id']
                    res = api.d_conv_info (id, conv['team_id'])
                    c.append (conv['id'])
                    if 'info' in res:
                        c.append (res['info'][0]['name'])
                    else: # d_conv_info () でチャンネルが not_found 等のエラーになった
                        # Todo: channel_not_found の扱いについて Slack に確認中
                        c.append (res['error'])
                    c.append (conv['team_id'])
                    c.append (teams[conv['team_id']][0])
                    c.append (teams[conv['team_id']][1])
                    c.append (c_type)
                    ret.append(c)

            with open(channels_csv, mode='a', encoding="utf_8_sig") as f:
                writer = csv.writer(f)
                for line in ret:
                    writer.writerow (line)

    print ("STEP2: get channels info target user has joind (include already left)... done")
    print ("  channels_csv -> ", channels_csv)


    """
    Step3: チャットログ出力用のディレクトリを作成する
    """
    # channels_csv から内容を取得し、teams のフォルダを作る
    # OrG も OrG の team フォルダを作る
    print ("STEP3: create directories for chat logs...", '\r', end='')
    target_teams = set()
    with open (channels_csv, mode='r') as f:
        reader = csv.reader(f)
        x = 0
        for elem in reader:
            x += 1
            print_progress_cycle ("STEP3: create directories for chat logs...", x)
            target_teams.add (elem[3])
        target_teams.remove('team_name') # csv の header 文字列
    for team in target_teams:
        x += 1
        print_progress_cycle ("STEP3: create directories for chat logs...", x)
        os.makedirs (OUTPUT_DIR + '/' + team)

    print ("STEP3: create directories for chat logs... done")

    """
    Step4: ログ json を収集する
    """
    print ("STEP4: cp log json files from EXPORT_DIR...", '\r', end='')
    dst_list = []
    with open (channels_csv, mode='r') as f:
        reader = csv.reader(f)
        for i,elem in enumerate(reader):
            if i == 0:
                pass
            else:
                print ("STEP4: cp log json files from EXPORT_DIR... " + ' : ' + elem[1], '\r', end='')
                if elem[4] == 'im' or elem[4] == "mpim": # OrG レベルオブジェクト
                    if not elem[1] == 'channel_not_found':
                        src = EXPORT_DIR + '/' + elem[1]
                        dst = OUTPUT_DIR + '/' + ORG_NAME + '/' + elem[1]
                        if os.path.exists(src):
                            shutil.copytree (src, dst)
                            dst_list.append(dst)
                    else:
                        pass
                        #print ("e4 channel_not_found", file=sys.stderr)
                else: # elem[4] == public or private
                    if elem[2] == ORG_ID: # OrG レベルのチャット
                        if not elem[1] == 'channel_not_found':
                            src = EXPORT_DIR + '/' + elem[1]
                            dst = OUTPUT_DIR + '/' + ORG_NAME + '/' + elem[1]
                            if os.path.exists (src):
                                shutil.copytree (src, dst)
                                dst_list.append(dst)
                            else:
                                pass
                                #print ("e1 not in export files", file=sys.stderr)
                        else:
                            pass
                            #print ("e2 channel_not_found", file=sys.stderr)
                    else: # ワークスペース単位のチャット
                        src = EXPORT_DIR + '/teams/' + teams[elem[2]][0] + '-' + teams[elem[2]][1] + '/' + elem[1]
                        if os.path.exists(src):
                            dst = OUTPUT_DIR + '/' + elem[3] + '/' + elem[1]
                            shutil.copytree (src, dst)
                            dst_list.append(dst)
                        else:
                            pass
                            #print ("e3 not in export files", file=sys.stderr)
                print ("STEP4: cp log json files from EXPORT_DIR...                                                           ", '\r', end='') # 長さは適当

    print ("STEP4: cp log json files from EXPORT_DIR...                                                           ", '\r', end='') # 長さは適当
    print ("STEP4: cp log json files from EXPORT_DIR... done")


    """
    Step5: 集めた ログ.json を加工する
    """
    print ("STEP5: processing json files change into xlsx files... ", '\r', end='')
    ROW = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    HEADER = ['datetime(JST)', 'type', 'subtype', 'user_id', 'user_name', 'thread', 'text', 'files']
    fill = openpyxl.styles.PatternFill (patternType='solid', fgColor='ffff00')
    total = str(len(dst_list))

    for c, dst_dir in enumerate(dst_list):
        dst_path = dst_dir.split('/')
        print ("STEP5: processing json files change into xlsx files... " + str(c+1) + '/' + total, '\r', end='')
        # dst_dir にある *.json をリスト化する
        # 一つのディレクトリに1つのxlsxファイル、一つのjsonファイルを一つのワークシートにまとめる
        # wb 単位の処理
        json_files = glob.glob (dst_dir + '/*.json')
        wb_path = 'templete.xlsx'
        output_wb_name = dst_dir.split('/')[-1]
        output_wb_path = dst_dir + '/' + output_wb_name + '.xlsx'
        wb = openpyxl.load_workbook(wb_path)
        for json_file in json_files:
            # ws 単位の処理
            ws_name = os.path.splitext(os.path.basename(json_file))[0]
            ws = wb.copy_worksheet (wb['Sheet1'])
            ws.title = ws_name
            for r,h in zip (ROW, HEADER):
                ws[r+'1'] = h

            with open (json_file, mode='r') as f:
                j = json.load (f)

            r = 1
            for elem in j:

                dt_jst = datetime.datetime.fromtimestamp (float(elem['ts']), datetime.timezone(datetime.timedelta(hours=9)))
                elem_type = elem['type']
                elem_subtype = elem['subtype'] if 'subtype' in elem else '-'

                # user id #
                elem_user = '-'
                if 'user' in elem:
                    elem_user = elem['user'] 
                elif 'message' in elem:
                    if 'user' in elem['message']:
                        elem_user = elem['message']['user'] 

                # user name #
                elem_username = '-'
                if 'user_profile' in elem:
                    if 'display_name' in elem['user_profile']:
                        elem_username = elem['user_profile']['display_name']  
                elif 'root' in elem:
                    if 'user_profile' in elem['root']:
                        if 'display_name' in elem['root']['user_profile']:
                            elem_username = elem['root']['user_profile']['display_name']

                # thread_ts #
                elem_threadts = '-'
                if 'thread_ts' in elem:
                    elem_threadts = elem['thread_ts']
                elif 'message' in elem:
                    if 'thread_ts' in elem['message']:
                        elem_threadts = elem['message']['thread_ts']

                # text #
                elem_text = '-'
                if elem_subtype == 'message_changed':
                    if 'original' in elem:
                        elem_text = elem['original']['text']
                    elif 'message' in elem:  # ※ message_changed だが、 original がない、どう change したのか不明（text内容は変化ない？）
#                       elem_text = elem['message']['text']
                        continue # この場合の処理はいったん不要とする
                    else:
                        #print ('e9', file=sys.stderr)
                        exit ()
                else:
                    if 'text' in elem:
                        elem_text = elem['text']
                    elif 'message' in elem:
                        if 'text' in elem['message']:
                            elem_text = elem['message']['text']

                # files #
                files = ''
                if 'files' in elem:
                    if files:
                        for i in range(len(elem['files'])):
                            if 'name' in elem['files'][i]:
                                files += elem['files'][i]['name'] + '\n'
                            elif 'mode' in elem['files'][i]:
                                files += elem['files'][i]['mode'] + '\n'
                            else: # 上記以外の例外
                                #print ('e7', file=sys.stderr)
                                pass
                elif 'root' in elem:
                    if 'files' in elem['root']:
                        for i in range(len(elem['root']['files'])):
                            if 'name' in elem['root']['files'][i]:
                                files += elem['root']['files'][i]['name'] + '\n'
                            elif 'mode' in elem['root']['files'][i]:
                                files += elem['root']['files'][i]['mode'] + '\n'
                            else: # 上記以外の例外
                                #print ('e7', file=sys.stderr)
                                pass

                msg = [f"{dt_jst:%Y-%m-%d %H:%M:%S}", elem_type, elem_subtype, elem_user, elem_username, elem_threadts, elem_text, files]

                r += 1 # 行を挿入する直前でインクリメントする
                for x, m in zip (ROW, msg):
                    try:
                        ws[x+str(r)] = m
                    except openpyxl.utils.exceptions.IllegalCharacterError:
                        # openpyxl で処理できない文字列を検出した場合
                        # ref: https://qiita.com/que9/items/8326133721500741bde7
                        m = "Sanitized openpyxl.utils.exceptions.IllegalCharacterError"
                        ws[x+str(r)] = m
                    if elem_user == user:
                        ws[x+str(r)].fill = fill

        ws = wb['Sheet1']
        wb.remove (ws)
        if not len(wb.sheetnames) == 0:
            wb.save (output_wb_path)

    print ("STEP5: processing json files change into xlsx files... done             ") # 長さは適当
    print ("  output_dir -> " + OUTPUT_DIR)

if __name__ == '__main__':
    main()
    exit (0)
